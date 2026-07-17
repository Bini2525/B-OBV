"""
OBV Trend- & Divergenzanalyse
==============================

Streamlit-App zur automatisierten On-Balance-Volume (OBV) Analyse von Aktien.

Funktionen:
    - Ticker-Eingabe manuell (Textfeld, kommagetrennt, inkl. WKN für deutsche
      Werte) ODER per Excel-Upload (inkl. WKN-Spaltenerkennung)
    - Automatische Auflösung europäischer Ticker/WKNs ohne Börsen-Suffix
      (z. B. "BMW" -> "BMW.DE") über Suffix-Heuristik + Yahoo-Finance-Suche
    - Analyse-Zeitraum wählbar: 1 Woche, 1 Monat, 3 Monate, 6 Monate, 1 Jahr
    - Abruf historischer Kurs-/Volumendaten via yfinance
    - Berechnung des On-Balance-Volume (OBV)
    - Trend- und Divergenzanalyse (Kurs vs. OBV) über die letzten Handelstage
    - Durchschnittliches Analystenkursziel je Ticker (sofern von Yahoo Finance geführt)
    - Vergleich aktueller Kurs vs. Analystenkursziel (Über-/Unterbewertung in %)
    - Alle Kurse einheitlich in Euro (Live-Umrechnung über Yahoo-FX-Kurse)
    - Interaktive Ergebnistabelle in der App (deutsches Zahlenformat)
    - Download des Ergebnisses als formatierte Excel-Datei (.xlsx)

Hinweis: Diese App dient ausschließlich Informationszwecken und stellt
keine Anlageberatung dar.

Hosting:
    Kann unverändert über GitHub + Streamlit Community Cloud deployed werden.
    Einstiegsdatei: app.py | Abhängigkeiten: requirements.txt
"""

from __future__ import annotations

import io
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Konstanten / Konfiguration
# ---------------------------------------------------------------------------

# Zuordnung der UI-Auswahl auf die von yfinance erwarteten Period-Strings.
# "1 Woche" wird über "5d" (5 Handelstage) abgebildet, da yfinance keinen
# eigenen 1-Wochen-Period-String kennt.
PERIOD_OPTIONS: dict[str, str] = {
    "1 Woche": "5d",
    "1 Monat": "1mo",
    "3 Monate": "3mo",
    "6 Monate": "6mo",
    "1 Jahr": "1y",
}
DEFAULT_PERIOD_LABEL = "6 Monate"

# Mindestanzahl an Handelstagen, ab der die Trend-/Divergenzanalyse als
# statistisch belastbar gilt. Bei kürzeren Zeiträumen (z. B. "1 Woche")
# wird das Ergebnis trotzdem berechnet, aber als eingeschränkt gekennzeichnet.
MIN_RELIABLE_DAYS = 10

# Anzahl der Handelstage, die für die Trend-/Divergenzanalyse betrachtet werden.
LOOKBACK_DAYS = 15  # liegt im geforderten Korridor von 10 bis 20 Handelstagen

# Schwellenwert für die Trendklassifikation (relative Steigung pro Tag).
# Ein Wert von 0.001 entspricht einer durchschnittlichen Veränderung von
# ca. 0,1 % pro Handelstag und dient dazu, "Rauschen" von einem echten
# Trend zu unterscheiden.
TREND_SLOPE_THRESHOLD = 0.001

# Spaltennamen, nach denen beim Excel-Import gesucht wird (case-insensitive).
# "wkn" ergänzt, damit Broker-Exports mit Wertpapierkennnummer statt Ticker
# (z. B. von Trade Republic/comdirect) automatisch erkannt werden.
TICKER_COLUMN_CANDIDATES = ["ticker", "symbol", "aktie", "wertpapier", "wkn", "wertpapierkennnummer"]

# Pflicht-Disclaimer für alle Aktien-/Finanzanalyse-Apps.
DISCLAIMER_TEXT = (
    "Hinweis: Diese App dient ausschließlich Informationszwecken und stellt "
    "keine Anlageberatung dar. Keine Kauf- oder Verkaufsempfehlung."
)

# Gängige Börsen-Suffixe für Yahoo Finance, nach denen gesucht wird, wenn ein
# roher Ticker (z. B. "BMW" statt "BMW.DE") nicht direkt gefunden wird.
# Schwerpunkt DACH/Westeuropa, da für europäische Broker-Exports typisch.
EXCHANGE_SUFFIXES = [
    ".DE",  # Xetra (Deutschland)
    ".F",   # Frankfurt (Alternative zu Xetra)
    ".SW",  # SIX Swiss Exchange
    ".L",   # London Stock Exchange
    ".AS",  # Euronext Amsterdam
    ".PA",  # Euronext Paris
    ".MI",  # Borsa Italiana Mailand
    ".MC",  # Bolsa de Madrid
    ".BR",  # Euronext Brüssel
    ".VI",  # Wiener Börse
]

# Yahoo Finance meldet Kurse von London-notierten Werten oft in Pence statt
# Pfund (Währungscode "GBp"/"GBX"). Für die EUR-Umrechnung muss das vorher
# durch 100 geteilt werden, sonst wäre der Wert um Faktor 100 zu hoch.
PENCE_CURRENCIES = {"GBp", "GBX"}

# Cache für bereits abgerufene Wechselkurse innerhalb einer App-Sitzung, damit
# nicht pro Ticker erneut derselbe FX-Kurs abgerufen werden muss.
_FX_RATE_CACHE: dict[str, float | None] = {}

# Finale Spaltenreihenfolge für Anzeige & Excel-Export: zuerst Aktienname,
# dann Trend-Bestätigung/Divergenz, dann Differenz zum Analystenkursziel -
# das sind die drei Spalten, die für die schnelle Einschätzung am wichtigsten
# sind. Alle übrigen (eher nachrangigen) Spalten folgen danach.
COLUMN_ORDER = [
    "Name",
    "Trend-Bestätigung / Divergenz",
    "Kurs über oder unter Analystenziel",
    "Ticker",
    "Ursprüngliche Währung",
    "Aktueller Kurs (EUR)",
    "Durchschnittliches Analystenkursziel (EUR)",
    "Letztes Volumen",
    "Aktueller OBV-Wert",
]

# Sortierreihenfolge der Trend-/Divergenz-Kategorien für die Ausgabe (primäres
# Sortierkriterium). "Keine klare Richtung" ist absichtlich nicht enthalten
# und landet über den Fallback-Wert TREND_SORT_FALLBACK am Ende der Liste.
TREND_SORT_ORDER = {
    "Aufwärtstrend bestätigt": 0,
    "Bullische Divergenz (Kaufsignal)": 1,
    "Abwärtstrend bestätigt": 2,
    "Bärische Divergenz (Verkaufssignal)": 3,
}
TREND_SORT_FALLBACK = len(TREND_SORT_ORDER)  # z. B. für "Keine klare Richtung"


# ---------------------------------------------------------------------------
# Hilfsfunktionen: Ticker-Eingabe (manuell & Excel-Upload)
# ---------------------------------------------------------------------------

def parse_manual_tickers(raw_text: str) -> list[str]:
    """Wandelt eine kommagetrennte Texteingabe in eine bereinigte Ticker-Liste um."""
    if not raw_text:
        return []
    tickers = [t.strip().upper() for t in raw_text.split(",")]
    return [t for t in tickers if t]  # leere Einträge entfernen


def extract_tickers_from_excel(uploaded_file) -> list[str]:
    """
    Liest eine hochgeladene Excel-Datei ein und extrahiert die Ticker-Spalte.

    Sucht case-insensitive nach gängigen Spaltennamen (ticker, symbol, aktie,
    wertpapier, wkn). Wird keine passende Spalte gefunden, wird die erste
    Spalte der Tabelle verwendet.
    """
    df = pd.read_excel(uploaded_file)

    if df.empty or df.shape[1] == 0:
        return []

    # Case-insensitive Suche nach einer passenden Spaltenbezeichnung.
    ticker_column = None
    for col in df.columns:
        if str(col).strip().lower() in TICKER_COLUMN_CANDIDATES:
            ticker_column = col
            break

    # Fallback: erste Spalte der Tabelle verwenden.
    if ticker_column is None:
        ticker_column = df.columns[0]

    raw_values = df[ticker_column].dropna().astype(str).tolist()
    tickers = [v.strip().upper() for v in raw_values]
    return [t for t in tickers if t]


def get_ticker_list(input_mode: str, manual_text: str, uploaded_file) -> list[str]:
    """Ermittelt je nach gewähltem Eingabemodus die finale, deduplizierte Ticker-Liste."""
    if input_mode == "Manuelle Eingabe":
        tickers = parse_manual_tickers(manual_text)
    else:
        tickers = extract_tickers_from_excel(uploaded_file)

    # Reihenfolge beibehalten, aber Duplikate entfernen.
    seen = set()
    unique_tickers = []
    for t in tickers:
        if t not in seen:
            seen.add(t)
            unique_tickers.append(t)
    return unique_tickers


# ---------------------------------------------------------------------------
# Ticker-Auflösung (Ticker, WKN, ISIN -> Yahoo-Symbol) & Datenabruf
# ---------------------------------------------------------------------------

def _try_symbol(symbol: str, period: str):
    """
    Versucht, für ein konkretes Yahoo-Finance-Symbol Kursdaten zu laden.

    Gibt (ticker_obj, hist) zurück, wenn verwertbare Daten vorliegen,
    sonst None. Wirft absichtlich keine Exception - wird als "Testballon"
    innerhalb der Auflösungs-Kette verwendet.
    """
    try:
        ticker_obj = yf.Ticker(symbol)
        hist = ticker_obj.history(period=period, auto_adjust=False)
        if hist is None or hist.empty:
            return None
        if "Close" not in hist.columns or "Volume" not in hist.columns:
            return None
        hist = hist[["Close", "Volume"]].dropna()
        if hist.empty:
            return None
        return ticker_obj, hist
    except Exception:
        return None


def _search_candidates(query: str, period: str):
    """
    Fragt die Yahoo-Finance-Volltextsuche ab und liefert (symbol, ticker_obj,
    hist) für den ersten Treffer mit verwertbaren Kursdaten zurück, sonst
    None. Nimmt in einem ersten Durchlauf nur Treffer vom Typ "EQUITY" (echte
    Aktien statt z. B. News/Optionen), im zweiten Durchlauf alle Treffer -
    das verbessert die Trefferqualität bei mehrdeutigen Suchbegriffen wie
    einer WKN.
    """
    try:
        quotes = yf.Search(query, max_results=8).quotes
    except Exception:
        quotes = []

    if not quotes:
        return None

    equity_quotes = [q for q in quotes if q.get("quoteType") == "EQUITY"]
    other_quotes = [q for q in quotes if q.get("quoteType") != "EQUITY"]

    for quote in equity_quotes + other_quotes:
        symbol = quote.get("symbol")
        if not symbol:
            continue
        found = _try_symbol(symbol, period)
        if found is not None:
            return symbol, found[0], found[1]

    return None


def resolve_ticker(raw_ticker: str, period: str) -> tuple[str, "yf.Ticker", pd.DataFrame]:
    """
    Löst einen rohen Ticker-Code, eine WKN oder eine ISIN in ein gültiges
    Yahoo-Finance-Symbol auf und lädt dabei direkt die passende Kurshistorie.

    Viele europäische Broker-Exports (z. B. Trade Republic, comdirect) führen
    Aktien ohne Börsen-Suffix (z. B. "BMW" statt "BMW.DE") oder nur mit WKN
    (z. B. "716460" für BMW). Auflösungs-Reihenfolge:

        1. Eingabe wie angegeben probieren (deckt US-Ticker sowie bereits
           korrekt angegebene Symbole wie "BMW.DE" ab).
        2. Yahoo-Finance-Volltextsuche (yfinance.Search) - findet in der
           Praxis auch WKN und ISIN, da Yahoo diese für europäische Werte
           mitindiziert, und liefert direkt das korrekte Börsensymbol.
        3. Als letzter Fallback: gängige europäische Börsensuffixe (siehe
           EXCHANGE_SUFFIXES) systematisch durchprobieren.

    Gibt (aufgelöstes_symbol, ticker_obj, hist) zurück oder wirft eine
    ValueError, wenn keiner der drei Wege zu verwertbaren Daten führt.
    """
    # 1) Eingabe wie angegeben (Ticker oder bereits korrektes Yahoo-Symbol).
    found = _try_symbol(raw_ticker, period)
    if found is not None:
        return raw_ticker, found[0], found[1]

    # 2) Yahoo-Finance-Suche - deckt Ticker, WKN und ISIN ab.
    search_result = _search_candidates(raw_ticker, period)
    if search_result is not None:
        return search_result

    # 3) Suffix-Heuristik als letzter Fallback (nur wenn kein Suffix bereits
    #    im Ticker enthalten ist - sonst würden unsinnige Kombinationen wie
    #    "BMW.DE.SW" entstehen). Für reine WKN-Eingaben (nur Ziffern/Buchstaben
    #    ohne Punkt) ebenfalls sinnvoll, falls die Suche nichts fand.
    if "." not in raw_ticker:
        for suffix in EXCHANGE_SUFFIXES:
            candidate = f"{raw_ticker}{suffix}"
            found = _try_symbol(candidate, period)
            if found is not None:
                return candidate, found[0], found[1]

    raise ValueError(
        f"Kein gültiges Yahoo-Finance-Symbol für '{raw_ticker}' gefunden - "
        f"auch nicht über die Yahoo-Suche (Ticker/WKN/ISIN) oder gängige "
        f"Börsensuffixe (.DE/.SW/.L/...). Ticker ggf. mit explizitem "
        f"Börsenkürzel angeben, z. B. '{raw_ticker}.DE'."
    )


def get_fx_rate_to_eur(currency: str) -> float | None:
    """
    Ermittelt den aktuellen Wechselkurs von `currency` nach EUR über den
    Yahoo-Finance-FX-Ticker "{currency}EUR=X" (Konvention: "AAABBB=X" zeigt
    an, wie viele Einheiten BBB man für 1 Einheit AAA erhält - "USDEUR=X"
    liefert also den EUR-Gegenwert von 1 USD).

    Nutzt einen einfachen In-Memory-Cache pro Sitzung, damit derselbe Kurs
    nicht für jeden Ticker erneut abgerufen wird. Gibt None zurück, wenn der
    Kurs nicht ermittelt werden konnte (z. B. FX-Paar nicht verfügbar).
    """
    if currency in _FX_RATE_CACHE:
        return _FX_RATE_CACHE[currency]

    rate: float | None = None
    try:
        fx_ticker = yf.Ticker(f"{currency}EUR=X")
        hist = fx_ticker.history(period="5d")
        if hist is not None and not hist.empty and "Close" in hist.columns:
            close_values = hist["Close"].dropna()
            if not close_values.empty:
                rate = float(close_values.iloc[-1])
    except Exception:
        rate = None

    _FX_RATE_CACHE[currency] = rate
    return rate


def convert_to_eur(value: float | None, currency: str) -> float | None:
    """
    Rechnet einen Kurswert von der Handelswährung in Euro um.

    Behandelt den Sonderfall britischer Pence (Währungscode "GBp"/"GBX",
    von Yahoo für London-notierte Werte verwendet) korrekt, indem zunächst
    durch 100 geteilt wird (Pence -> Pfund), bevor der GBP/EUR-Kurs
    angewendet wird.

    Gibt None zurück, wenn keine Umrechnung möglich ist (z. B. unbekannte
    Währung oder FX-Kurs nicht abrufbar) - der Aufrufer zeigt in diesem
    seltenen Fall den Originalwert unverändert an, statt abzustürzen.
    """
    if value is None:
        return None
    if not currency or currency == "n/a":
        return None

    normalized_currency = currency
    divisor = 1.0
    if currency in PENCE_CURRENCIES:
        normalized_currency = "GBP"
        divisor = 100.0

    if normalized_currency == "EUR":
        return round(value / divisor, 2)

    rate = get_fx_rate_to_eur(normalized_currency)
    if rate is None:
        return None

    return round((value / divisor) * rate, 2)


def fetch_price_history(ticker: str, period: str) -> tuple[pd.DataFrame, dict]:
    """
    Ruft historische Kurs- und Volumendaten für einen Ticker ab (inkl.
    automatischer Ticker-/WKN-/ISIN-Auflösung, siehe resolve_ticker).

    Gibt ein DataFrame mit den Spalten 'Close' und 'Volume' sowie ein
    Meta-Dict (Name, Handelswährung, durchschnittliches Analystenkursziel,
    aufgelöstes Yahoo-Symbol) zurück. Wirft eine Exception, wenn keine
    verwertbaren Daten vorliegen (wird vom Aufrufer abgefangen).
    """
    resolved_symbol, ticker_obj, hist = resolve_ticker(ticker, period)

    meta = {
        "currency": "n/a",
        "avg_analyst_target": None,
        "name": resolved_symbol,
        "resolved_symbol": resolved_symbol,
    }

    # Firmenname, Handelswährung & durchschnittliches Analystenkursziel
    # ermitteln. Alles nicht kritisch für die OBV-Analyse selbst (nicht jeder
    # Ticker hat z. B. Analysten-Coverage), daher robust mit Fallback statt
    # Abbruch.
    try:
        info = ticker_obj.info
        if info:
            meta["currency"] = info.get("currency") or meta["currency"]
            target = info.get("targetMeanPrice")
            if target is not None:
                meta["avg_analyst_target"] = float(target)
            name = info.get("longName") or info.get("shortName")
            if name:
                meta["name"] = name
    except Exception:
        pass

    if meta["currency"] == "n/a":
        try:
            fast_info = getattr(ticker_obj, "fast_info", None)
            if fast_info is not None:
                meta["currency"] = fast_info.get("currency") or "n/a"
        except Exception:
            pass

    return hist, meta


def calculate_obv(hist: pd.DataFrame) -> pd.DataFrame:
    """
    Berechnet das On-Balance-Volume (OBV) für die übergebene Kurshistorie.

    Regeln:
        Close(t) > Close(t-1)  -> OBV(t) = OBV(t-1) + Volume(t)
        Close(t) < Close(t-1)  -> OBV(t) = OBV(t-1) - Volume(t)
        Close(t) == Close(t-1) -> OBV(t) = OBV(t-1)
    """
    df = hist.copy()
    close_diff = df["Close"].diff()

    # Richtung: +1 (Anstieg), -1 (Rückgang), 0 (unverändert).
    direction = np.sign(close_diff).fillna(0)

    # Signierte Volumina aufkumulieren -> OBV. Der erste Wert startet bei 0.
    signed_volume = direction * df["Volume"]
    df["OBV"] = signed_volume.cumsum()
    df["OBV"] = df["OBV"].fillna(0)

    return df


# ---------------------------------------------------------------------------
# Trend- & Divergenzanalyse
# ---------------------------------------------------------------------------

def classify_trend(series: pd.Series, threshold: float = TREND_SLOPE_THRESHOLD) -> str:
    """
    Klassifiziert den Trend einer Zeitreihe als 'up', 'down' oder 'flat'.

    Methode: lineare Regression (Steigung) über das Analysefenster,
    normalisiert auf das mittlere Niveau der Reihe, damit die Klassifikation
    unabhängig von der absoluten Größenordnung (Kurs vs. OBV) funktioniert.
    """
    series = series.dropna()
    if len(series) < 2:
        return "flat"

    x = np.arange(len(series))
    slope = np.polyfit(x, series.values, 1)[0]

    avg_level = np.mean(np.abs(series.values))
    if avg_level == 0:
        avg_level = 1.0  # Division durch 0 vermeiden

    normalized_slope = slope / avg_level

    if normalized_slope > threshold:
        return "up"
    if normalized_slope < -threshold:
        return "down"
    return "flat"


def analyze_trend_divergence(df: pd.DataFrame, lookback: int = LOOKBACK_DAYS) -> str:
    """
    Vergleicht den Kurstrend mit dem OBV-Trend über die letzten `lookback`
    Handelstage und leitet daraus eine Bewertung ab.

    Regeln:
        Kurs up   + OBV up            -> Aufwärtstrend bestätigt
        Kurs down + OBV down          -> Abwärtstrend bestätigt
        Kurs down/flat + OBV up       -> Bullische Divergenz (Kaufsignal)
        Kurs up + OBV down/flat       -> Bärische Divergenz (Verkaufssignal)
        alle anderen Kombinationen    -> Keine klare Richtung
    """
    window = df.tail(lookback)

    price_trend = classify_trend(window["Close"])
    obv_trend = classify_trend(window["OBV"])

    if price_trend == "up" and obv_trend == "up":
        bewertung = "Aufwärtstrend bestätigt"
    elif price_trend == "down" and obv_trend == "down":
        bewertung = "Abwärtstrend bestätigt"
    elif price_trend in ("down", "flat") and obv_trend == "up":
        bewertung = "Bullische Divergenz (Kaufsignal)"
    elif price_trend == "up" and obv_trend in ("down", "flat"):
        bewertung = "Bärische Divergenz (Verkaufssignal)"
    else:
        bewertung = "Keine klare Richtung"

    # Bei kurzen Zeiträumen (z. B. "1 Woche") steht weniger als
    # MIN_RELIABLE_DAYS Handelstage zur Verfügung - Ergebnis wird trotzdem
    # ausgegeben, aber transparent als weniger belastbar gekennzeichnet.
    if len(window) < MIN_RELIABLE_DAYS:
        bewertung += " (eingeschränkte Datenbasis)"

    return bewertung


# ---------------------------------------------------------------------------
# Deutsches Zahlenformat (wird bereits hier gebraucht: compare_price_to_target)
# ---------------------------------------------------------------------------

def format_de_number(value, decimals: int = 2) -> str:
    """
    Formatiert eine Zahl im deutschen Format (Tausenderpunkt, Komma als
    Dezimaltrennzeichen), z. B. 1234.5 -> "1.234,50".

    Gibt einen leeren String zurück, wenn kein Wert vorhanden ist
    (None/NaN) - relevant z. B. für fehlendes Analystenkursziel.
    """
    if value is None:
        return ""
    try:
        if isinstance(value, float) and np.isnan(value):
            return ""
    except TypeError:
        pass

    # Trick: zunächst im US-Format formatieren (Komma=Tausender, Punkt=Dezimal),
    # anschließend Trennzeichen tauschen -> deutsches Format.
    us_formatted = f"{value:,.{decimals}f}"
    return us_formatted.replace(",", "§").replace(".", ",").replace("§", ".")


# ---------------------------------------------------------------------------
# Vergleich: aktueller Kurs vs. durchschnittliches Analystenkursziel
# ---------------------------------------------------------------------------

def compute_diff_pct(current_price: float | None, avg_target: float | None) -> float | None:
    """
    Berechnet die vorzeichenbehaftete prozentuale Abweichung des aktuellen
    Kurses vom durchschnittlichen Analystenkursziel: negativ = Kurs liegt
    UNTER dem Kursziel (potenzielles Aufwärtspotenzial), positiv = Kurs
    liegt ÜBER dem Kursziel.

    Wird sowohl für den Anzeigetext (compare_price_to_target) als auch für
    die Sortierung der Ergebnistabelle (siehe main()) verwendet, damit beide
    Stellen garantiert denselben Wert zugrunde legen.

    Gibt None zurück, wenn keine Berechnung möglich ist (fehlender Kurs,
    fehlendes Kursziel oder Kursziel = 0).
    """
    if current_price is None or avg_target is None or avg_target == 0:
        return None
    return (current_price - avg_target) / avg_target * 100


def compare_price_to_target(current_price: float | None, avg_target: float | None) -> str:
    """
    Vergleicht den aktuellen Kurs mit dem durchschnittlichen Analystenkursziel
    (beide bereits in Euro umgerechnet) und gibt einen lesbaren Text mit
    prozentualer Abweichung zurück, z. B. "12,34 % über Kursziel" oder
    "8,50 % unter Kursziel".

    Gibt "keine Analysten-Coverage" zurück, wenn kein Analystenkursziel oder
    kein umgerechneter Kurs vorliegt, oder das Kursziel 0 ist (Division durch
    0 vermeiden). Bewusst NICHT "n/a" verwendet: pandas/Excel-Re-Importe
    interpretieren "n/a" oft automatisch als fehlenden Wert (NaN) statt als
    Text.
    """
    diff_pct = compute_diff_pct(current_price, avg_target)
    if diff_pct is None:
        return "keine Analysten-Coverage"

    diff_pct_str = format_de_number(abs(diff_pct), 2)

    if diff_pct > 0.005:
        return f"{diff_pct_str} % über Kursziel"
    if diff_pct < -0.005:
        return f"{diff_pct_str} % unter Kursziel"
    return "genau auf Kursziel"


def trend_sort_key(bewertung: str) -> int:
    """
    Ordnet einen Trend-/Divergenz-Text seiner Sortierposition gemäß
    TREND_SORT_ORDER zu (Reihenfolge: Aufwärtstrend bestätigt, Bullische
    Divergenz, Abwärtstrend bestätigt, Bärische Divergenz).

    Der Zusatz " (eingeschränkte Datenbasis)" (siehe analyze_trend_divergence)
    wird vor dem Abgleich entfernt, damit auch eingeschränkt belastbare
    Ergebnisse korrekt einsortiert werden. Alles andere (z. B. "Keine klare
    Richtung") landet über TREND_SORT_FALLBACK am Ende.
    """
    base = bewertung.replace(" (eingeschränkte Datenbasis)", "")
    return TREND_SORT_ORDER.get(base, TREND_SORT_FALLBACK)


# ---------------------------------------------------------------------------
# Analyse-Pipeline pro Ticker
# ---------------------------------------------------------------------------

def analyze_ticker(ticker: str, period: str) -> tuple[dict | None, str | None]:
    """
    Führt die vollständige Analyse für einen einzelnen Ticker aus.

    Gibt ein Tupel (Ergebnis-Dict, Fehlermeldung) zurück. Genau eines der
    beiden Elemente ist None - so kann der Aufrufer sauber zwischen Erfolg
    und Fehlschlag unterscheiden, ohne die App abstürzen zu lassen.
    """
    try:
        hist, meta = fetch_price_history(ticker, period)
        hist_with_obv = calculate_obv(hist)

        if len(hist_with_obv) < 2:
            return None, f"{ticker}: Zu wenige Datenpunkte für eine Analyse."

        bewertung = analyze_trend_divergence(hist_with_obv, LOOKBACK_DAYS)

        native_currency = meta.get("currency", "n/a")
        native_price = float(hist_with_obv["Close"].iloc[-1])
        native_target = meta.get("avg_analyst_target")

        # Alle Kurse einheitlich in Euro anzeigen (Live-Wechselkurs). Falls
        # die Umrechnung ausnahmsweise nicht möglich ist (z. B. FX-Paar bei
        # Yahoo nicht verfügbar), wird als Fallback der Originalwert in der
        # Handelswährung übernommen, damit kein Wert verloren geht.
        eur_price = convert_to_eur(native_price, native_currency)
        if eur_price is None:
            eur_price = round(native_price, 2)
        eur_target = convert_to_eur(native_target, native_currency) if native_target is not None else None

        result = {
            "Ticker": ticker,
            "Name": meta.get("name") or ticker,
            "Ursprüngliche Währung": native_currency,
            "Aktueller Kurs (EUR)": eur_price,
            "Durchschnittliches Analystenkursziel (EUR)": eur_target,
            "Kurs über oder unter Analystenziel": compare_price_to_target(eur_price, eur_target),
            "Letztes Volumen": int(hist_with_obv["Volume"].iloc[-1]),
            "Aktueller OBV-Wert": int(hist_with_obv["OBV"].iloc[-1]),
            "Trend-Bestätigung / Divergenz": bewertung,
            # Interne Sortier-Hilfsspalten (werden vor Anzeige/Export wieder
            # entfernt, siehe sort_result_dataframe in main()):
            "_trend_sort": trend_sort_key(bewertung),
            "_diff_pct_sort": compute_diff_pct(eur_price, eur_target),
        }
        return result, None

    except Exception as exc:  # noqa: BLE001 - bewusst breit, damit kein Ticker die App killt
        return None, f"{ticker}: Analyse fehlgeschlagen ({exc})."


def sort_and_order_results(result_df: pd.DataFrame) -> pd.DataFrame:
    """
    Sortiert die Ergebniszeilen und bringt die Spalten in die gewünschte
    Endreihenfolge (COLUMN_ORDER).

    Sortierung (zwei Ebenen):
        1. Trend-/Divergenz-Kategorie in der Reihenfolge Aufwärtstrend
           bestätigt -> Bullische Divergenz -> Abwärtstrend bestätigt ->
           Bärische Divergenz -> alles andere (z. B. "Keine klare Richtung").
        2. Innerhalb jeder Kategorie aufsteigend nach der vorzeichenbehafteten
           Differenz zum Analystenkursziel (_diff_pct_sort): das sortiert
           zunächst die am stärksten UNTER dem Kursziel liegenden Werte nach
           vorn (größte Unterbewertung zuerst), dann Richtung 0 (Kurs = Ziel),
           danach die ÜBER dem Kursziel liegenden Werte von der kleinsten bis
           zur größten Überbewertung. Ticker ohne Analysten-Coverage (kein
           Kursziel, _diff_pct_sort = NaN) werden dabei automatisch ans Ende
           der jeweiligen Kategorie sortiert (pandas-Standardverhalten für
           NaN bei aufsteigender Sortierung).

    Die beiden internen Sortier-Hilfsspalten werden danach entfernt, sie
    sollen weder in der Bildschirmanzeige noch im Excel-Export auftauchen.
    """
    sorted_df = result_df.sort_values(
        by=["_trend_sort", "_diff_pct_sort"],
        ascending=[True, True],
        na_position="last",
    ).reset_index(drop=True)

    sorted_df = sorted_df.drop(columns=["_trend_sort", "_diff_pct_sort"])

    # Spalten in die gewünschte Endreihenfolge bringen. Falls künftig eine
    # Spalte entfällt/hinzukommt, robust bleiben: nur vorhandene Spalten aus
    # COLUMN_ORDER übernehmen, danach eventuell verbleibende Spalten anhängen.
    ordered_columns = [c for c in COLUMN_ORDER if c in sorted_df.columns]
    remaining_columns = [c for c in sorted_df.columns if c not in ordered_columns]
    return sorted_df[ordered_columns + remaining_columns]


def build_display_dataframe(result_df: pd.DataFrame) -> pd.DataFrame:
    """
    Erstellt eine Anzeige-Kopie des Ergebnis-DataFrames mit Zahlen im
    deutschen Format - ausschließlich für die Bildschirmanzeige in
    st.dataframe(). Der Excel-Export (build_excel_bytes) arbeitet weiterhin
    mit den numerischen Rohwerten aus result_df, damit Excel eigene
    Zahlenformate korrekt anwenden kann.
    """
    display_df = result_df.copy()

    two_decimal_cols = ["Aktueller Kurs (EUR)", "Durchschnittliches Analystenkursziel (EUR)"]
    zero_decimal_cols = ["Letztes Volumen", "Aktueller OBV-Wert"]

    for col in two_decimal_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda v: format_de_number(v, 2))
    for col in zero_decimal_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda v: format_de_number(v, 0))

    return display_df


def _max_text_length(values, minimum: int = 0) -> int:
    """
    Ermittelt die maximale Textlänge einer Spalte, robust gegenüber None/NaN
    und unabhängig vom pandas-Backend (numpy-object vs. Arrow-backed).

    Hintergrund: `series.astype(str).map(len)` ist NICHT sicher, wenn die
    Spalte None-Werte enthält - je nach pandas-/pyarrow-Version bleibt der
    Null-Wert nach astype(str) ein NA-Sentinel statt der Textstring "None"
    zu werden, wodurch len() mit TypeError abbricht (genau dieser Fehler
    trat in Produktion bei fehlendem Analystenkursziel auf). Deshalb hier
    bewusst eine explizite pd.notna()-Prüfung statt astype(str) + map(len).
    """
    lengths = [len(str(v)) for v in values if pd.notna(v)]
    return max(lengths, default=minimum)


# ---------------------------------------------------------------------------
# Excel-Export mit Formatierung
# ---------------------------------------------------------------------------

def build_excel_bytes(result_df: pd.DataFrame) -> bytes:
    """Erstellt eine formatierte .xlsx-Datei (Bytes) aus dem Ergebnis-DataFrame."""
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="OBV-Analyse")
        worksheet = writer.sheets["OBV-Analyse"]

        header_font = Font(name="Arial", size=11, bold=True)
        data_font = Font(name="Arial", size=11)
        data_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        thin_black = Side(border_style="thin", color="000000")
        table_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)

        # Textspalten: linksbündig. Alle übrigen Spalten (Zahlen) rechtsbündig,
        # 2 Nachkommastellen + Tausenderpunkt bzw. 0 Nachkommastellen gemäß
        # globalem Zahlenformat-Standard.
        text_columns = {
            "Ticker",
            "Name",
            "Ursprüngliche Währung",
            "Kurs über oder unter Analystenziel",
            "Trend-Bestätigung / Divergenz",
        }
        number_format_map = {
            "Aktueller Kurs (EUR)": "#,##0.00",
            "Durchschnittliches Analystenkursziel (EUR)": "#,##0.00",
            "Letztes Volumen": "#,##0",
            "Aktueller OBV-Wert": "#,##0",
        }

        n_rows, n_cols = result_df.shape

        for col_idx, col_name in enumerate(result_df.columns, start=1):
            col_letter = get_column_letter(col_idx)

            # Kopfzeile: fett, kein Zeilenumbruch, Arial 11.
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.font = header_font
            header_cell.alignment = Alignment(
                horizontal="left" if col_name in text_columns else "right",
                vertical="top",
                wrap_text=False,
            )
            header_cell.border = table_border

            # Datenzeilen: berechnete/nicht editierbare Werte -> blau hinterlegt.
            for row_idx in range(2, n_rows + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.fill = data_fill
                cell.border = table_border
                cell.alignment = Alignment(
                    horizontal="left" if col_name in text_columns else "right",
                    vertical="top",
                    wrap_text=True,
                )
                if col_name in number_format_map:
                    cell.number_format = number_format_map[col_name]

            # Spaltenbreite: Mindestbreite 13, sonst am längsten Inhalt
            # ausgerichtet, damit keine "####"-Anzeige entsteht. None/NaN
            # werden dabei übersprungen statt einen Fehler zu werfen (siehe
            # _max_text_length).
            max_content_len = max(
                len(str(col_name)),
                _max_text_length(result_df[col_name].tolist()) if n_rows > 0 else 0,
            )
            worksheet.column_dimensions[col_letter].width = max(13, max_content_len + 2)

        worksheet.freeze_panes = "A2"

        # Disclaimer unterhalb der Tabelle: keine Anlageberatung.
        disclaimer_row = n_rows + 3
        worksheet.cell(row=disclaimer_row, column=1, value=DISCLAIMER_TEXT)
        worksheet.cell(row=disclaimer_row, column=1).font = Font(name="Arial", size=9, italic=True)

    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

def main() -> None:
    st.set_page_config(page_title="OBV Trend- & Divergenzanalyse", layout="wide")
    st.title("📊 OBV Trend- & Divergenzanalyse")
    st.caption(
        "Berechnet das On-Balance-Volume (OBV) für beliebige Aktien und vergleicht "
        "den Kurstrend mit dem OBV-Trend, um Trendbestätigungen und Divergenzen zu "
        "erkennen. Alle Kurse werden einheitlich in Euro angezeigt (Live-Umrechnung "
        "über aktuelle Wechselkurse - die Ursprungswährung steht zur Transparenz in "
        "der Spalte 'Ursprüngliche Währung'; da der aktuelle statt des historischen "
        "Wechselkurses verwendet wird, ist die Umrechnung eine Näherung). Ticker ohne "
        "Börsenkürzel (z. B. 'BMW' statt 'BMW.DE') sowie WKN oder ISIN für deutsche "
        "Werte werden automatisch über die Yahoo-Finance-Suche und gängige "
        "europäische Börsensuffixe aufgelöst."
    )
    st.warning(f"⚠️ {DISCLAIMER_TEXT}")

    # --- Sidebar: Eingabe -------------------------------------------------
    st.sidebar.header("Eingabe")

    input_mode = st.sidebar.radio(
        "Wie möchtest du die Ticker eingeben?",
        options=["Manuelle Eingabe", "Excel-Datei hochladen"],
    )

    manual_text = ""
    uploaded_file = None

    if input_mode == "Manuelle Eingabe":
        manual_text = st.sidebar.text_input(
            "Ticker, WKN oder ISIN (einzeln oder kommagetrennt)",
            placeholder="z. B. AAPL oder AAPL, BMW.DE, 716460",
            help=(
                "Für deutsche Werte kann statt des Tickers auch die WKN "
                "(z. B. '716460' für BMW) oder die ISIN eingegeben werden."
            ),
        )
    else:
        uploaded_file = st.sidebar.file_uploader(
            "Excel-Datei mit Ticker-/WKN-Liste (.xlsx / .xls)",
            type=["xlsx", "xls"],
            help=(
                "Es wird automatisch nach einer Spalte 'Ticker', 'Symbol', 'Aktie', "
                "'Wertpapier' oder 'WKN' gesucht. Wird keine gefunden, wird die "
                "erste Spalte der Tabelle verwendet."
            ),
        )

    period_label = st.sidebar.selectbox(
        "Analyse-Zeitraum",
        options=list(PERIOD_OPTIONS.keys()),
        index=list(PERIOD_OPTIONS.keys()).index(DEFAULT_PERIOD_LABEL),
    )

    start_analysis = st.sidebar.button("Analyse starten", type="primary")

    # --- Hauptbereich: Analyse & Ausgabe -----------------------------------
    if not start_analysis:
        st.info("Ticker/WKN eingeben bzw. Excel-Datei hochladen, Zeitraum wählen und "
                 "auf 'Analyse starten' klicken.")
        return

    tickers = get_ticker_list(input_mode, manual_text, uploaded_file)

    if not tickers:
        st.warning(
            "Es wurden keine Ticker gefunden. Bitte prüfe deine Eingabe bzw. die "
            "hochgeladene Excel-Datei."
        )
        return

    period = PERIOD_OPTIONS[period_label]

    results: list[dict] = []
    warnings: list[str] = []

    progress_bar = st.progress(0.0, text="Analyse läuft ...")
    for i, ticker in enumerate(tickers, start=1):
        result, error = analyze_ticker(ticker, period)
        if result is not None:
            results.append(result)
        if error is not None:
            warnings.append(error)
        progress_bar.progress(i / len(tickers), text=f"Analysiere {ticker} ...")
    progress_bar.empty()

    # Fehlerhafte / delistete Ticker freundlich melden, App läuft dennoch weiter.
    if warnings:
        st.warning(
            "Folgende Ticker konnten nicht analysiert werden und wurden "
            "übersprungen:\n\n" + "\n".join(f"- {w}" for w in warnings)
        )

    if not results:
        st.error("Für keinen der eingegebenen Ticker konnten Daten ermittelt werden.")
        return

    result_df = pd.DataFrame(results)
    result_df = sort_and_order_results(result_df)

    st.subheader("Ergebnis")
    st.dataframe(build_display_dataframe(result_df), use_container_width=True, hide_index=True)

    # --- Excel-Export -------------------------------------------------------
    excel_bytes = build_excel_bytes(result_df)
    file_date = datetime.today().strftime("%d.%m.%y")

    st.download_button(
        label="📥 Ergebnis als Excel herunterladen",
        data=excel_bytes,
        file_name=f"{file_date}_OBV_Analyse.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
